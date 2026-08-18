"""
utils/excel_utils.py
---------------------
Central place for every Excel read/write operation used by the Data Driven
Testing Framework (DDTF). Keeping this logic in one module means the test
layer (tests/test_login.py) never touches openpyxl directly.
"""

from datetime import datetime
from openpyxl import load_workbook

# Column positions inside test_data.xlsx (1-indexed, matches openpyxl convention)
COL_TEST_ID = 1
COL_USERNAME = 2
COL_PASSWORD = 3
COL_DATE = 4
COL_TIME = 5
COL_TESTER = 6
COL_RESULT = 7

SHEET_NAME = "LoginTestData"


def read_test_data(file_path: str) -> list[dict]:
    """
    Reads every credential row from the Excel sheet and returns it as a
    list of dictionaries. This list is what feeds pytest's parametrize
    step, which is the core of the Data Driven Testing approach.

    Returns rows in the shape:
        {"row": <excel row number>, "test_id": ..., "username": ..., "password": ...}
    """
    workbook = load_workbook(file_path)
    sheet = workbook[SHEET_NAME]

    test_data = []
    row_index = 2  # row 1 is the header
    while sheet.cell(row=row_index, column=COL_TEST_ID).value:
        test_data.append(
            {
                "row": row_index,
                "test_id": sheet.cell(row=row_index, column=COL_TEST_ID).value,
                # openpyxl returns None for an intentionally empty cell (e.g. blank
                # username/password test cases) — normalise that to "" for Selenium.
                "username": sheet.cell(row=row_index, column=COL_USERNAME).value or "",
                "password": sheet.cell(row=row_index, column=COL_PASSWORD).value or "",
            }
        )
        row_index += 1

    workbook.close()
    return test_data


def write_test_result(file_path: str, row: int, tester_name: str, result: str) -> None:
    """
    Writes the outcome of a single test execution back into the same
    Excel file: current Date, current Time of Test, Name of Tester and
    Test Result (Pass/Fail).

    `row` is the exact Excel row this result belongs to, supplied by
    read_test_data() so we never write to the wrong record even if rows
    are re-ordered.
    """
    workbook = load_workbook(file_path)
    sheet = workbook[SHEET_NAME]

    now = datetime.now()
    sheet.cell(row=row, column=COL_DATE, value=now.strftime("%Y-%m-%d"))
    sheet.cell(row=row, column=COL_TIME, value=now.strftime("%H:%M:%S"))
    sheet.cell(row=row, column=COL_TESTER, value=tester_name)
    sheet.cell(row=row, column=COL_RESULT, value=result)

    workbook.save(file_path)
    workbook.close()
