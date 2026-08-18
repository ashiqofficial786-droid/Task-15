"""
create_test_data.py
--------------------
One-time helper script that builds test_data.xlsx — the data source consumed
by the Data Driven Testing Framework (DDTF) in tests/test_login.py.

Run once with:  python data/create_test_data.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 1. Workbook / worksheet setup
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "LoginTestData"

HEADERS = [
    "Test ID",
    "Username",
    "Password",
    "Date",
    "Time of Test",
    "Name of Tester",
    "Test Result",
]

HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
BODY_FONT = Font(name="Arial", size=11)
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # yellow -> user-editable
AUTO_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # green  -> filled by automation

for col_idx, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")

# ---------------------------------------------------------------------------
# 2. Test data — 5 Username/Password combinations
#    Row 2 is the ONE valid OrangeHRM demo credential (used to prove the
#    "successful login" path). Rows 3-6 are intentionally invalid so the
#    suite also exercises the "failed login" path, which is standard
#    practice for login test coverage.
# ---------------------------------------------------------------------------
TEST_ROWS = [
    ("TC_LOGIN_01", "Admin", "admin123"),        # valid credentials -> expected PASS
    ("TC_LOGIN_02", "Admin", "wrongPass1"),       # wrong password -> expected FAIL
    ("TC_LOGIN_03", "InvalidUser", "admin123"),   # wrong username -> expected FAIL
    ("TC_LOGIN_04", "Admin", ""),                 # blank password -> expected FAIL
    ("TC_LOGIN_05", "", "admin123"),              # blank username -> expected FAIL
]

for row_idx, (test_id, username, password) in enumerate(TEST_ROWS, start=2):
    ws.cell(row=row_idx, column=1, value=test_id).font = BODY_FONT
    ws.cell(row=row_idx, column=2, value=username).font = BODY_FONT
    ws.cell(row=row_idx, column=3, value=password).font = BODY_FONT

    # Date / Time / Tester / Result are left blank here — the automation
    # framework (utils/excel_utils.py) fills these in when the test runs.
    for col_idx in range(4, 8):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = BODY_FONT
        cell.fill = AUTO_FILL

    # Mark the editable input cells (Username / Password) so a human
    # filling this sheet in later knows which cells are theirs to edit.
    ws.cell(row=row_idx, column=2).fill = INPUT_FILL
    ws.cell(row=row_idx, column=3).fill = INPUT_FILL

# ---------------------------------------------------------------------------
# 3. Legend explaining the color coding (placed a couple of rows below data)
# ---------------------------------------------------------------------------
legend_row = len(TEST_ROWS) + 4
ws.cell(row=legend_row, column=1, value="Legend:").font = Font(name="Arial", bold=True)
ws.cell(row=legend_row + 1, column=1, value="Yellow = user-editable input (Username / Password)").fill = INPUT_FILL
ws.cell(row=legend_row + 1, column=1).font = BODY_FONT
ws.cell(row=legend_row + 2, column=1, value="Green = auto-filled by the test framework after each run").fill = AUTO_FILL
ws.cell(row=legend_row + 2, column=1).font = BODY_FONT

# ---------------------------------------------------------------------------
# 4. Column widths for readability
# ---------------------------------------------------------------------------
WIDTHS = [14, 16, 14, 14, 14, 18, 14]
for idx, width in enumerate(WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(idx)].width = width

wb.save("data/test_data.xlsx")
print("test_data.xlsx created successfully.")
